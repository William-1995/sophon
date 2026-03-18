#!/usr/bin/env python3
"""Convert Excel (.xlsx, .xls) to CSV. Writes to output_path.

Reads input via path or params, writes to output_path. Requires openpyxl for .xlsx
and xlrd for .xls.
"""
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def _resolve_path(params: dict, file_path: str) -> Path:
    p = Path(file_path)
    if p.is_absolute():
        return p
    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    user_id = str(params.get("user_id") or "default_user")
    if p.parts and p.parts[0] == user_id:
        p = Path(*p.parts[1:]) if len(p.parts) > 1 else Path(".")
    return workspace_root / p


def _ensure_in_workspace(workspace_root: Path, target: Path) -> bool:
    try:
        target.resolve().relative_to(workspace_root.resolve())
        return True
    except ValueError:
        return False


def _xlsx_to_csv(file_path: Path, sheet: str | int) -> tuple[list[dict], int]:
    """Read xlsx sheet and return (rows as dicts, header count)."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
    else:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0
    headers = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        data.append({headers[j]: str(cell) if cell is not None else "" for j, cell in enumerate(row) if j < len(headers)})
    return data, len(headers)


def _xls_to_csv(file_path: Path, sheet: str | int) -> tuple[list[dict], int]:
    """Read xls sheet and return (rows as dicts, header count)."""
    if not HAS_XLRD:
        raise RuntimeError("xlrd not installed. Run: pip install xlrd")
    wb = xlrd.open_workbook(str(file_path))
    if isinstance(sheet, int):
        ws = wb.sheet_by_index(sheet) if sheet < wb.nsheets else wb.sheet_by_index(0)
    else:
        try:
            ws = wb.sheet_by_name(sheet)
        except xlrd.biffh.XLRDError:
            ws = wb.sheet_by_index(0)
    headers = [str(ws.cell_value(0, j)) for j in range(ws.ncols)]
    data = []
    for i in range(1, ws.nrows):
        data.append({headers[j]: str(ws.cell_value(i, j)) for j in range(ws.ncols)})
    return data, len(headers)


def main() -> None:
    params = json.load(sys.stdin)
    args = params.get("arguments") or params
    file_path = (args.get("file") or args.get("path") or "").strip()
    output_path = (args.get("output_path") or "").strip() or None
    sheet = args.get("sheet", 0)

    if not file_path:
        print(json.dumps({"error": "file (or path) is required"}))
        return
    if not output_path:
        print(json.dumps({"error": "output_path is required to write CSV"}))
        return

    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    full_path = _resolve_path(params, file_path)
    out_target = (workspace_root / output_path).resolve()

    if not full_path.exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        return
    if not _ensure_in_workspace(workspace_root, out_target):
        print(json.dumps({"error": "output_path cannot escape workspace"}))
        return

    suffix = full_path.suffix.lower()

    try:
        if suffix == ".csv":
            print(json.dumps({"error": "Input is already CSV; use filesystem to copy if needed"}))
            return
        if suffix in [".xlsx", ".xls"]:
            if suffix == ".xlsx":
                data, _ = _xlsx_to_csv(full_path, sheet)
            else:
                data, _ = _xls_to_csv(full_path, sheet)
            if not data:
                out_target.parent.mkdir(parents=True, exist_ok=True)
                out_target.write_text("", encoding="utf-8-sig")
                print(json.dumps({"output_path": output_path, "written": True, "rows": 0}))
                return
            headers = list(data[0].keys())
            out_target.parent.mkdir(parents=True, exist_ok=True)
            with open(out_target, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)
            print(json.dumps({"output_path": output_path, "written": True, "rows": len(data)}))
        else:
            print(json.dumps({"error": f"Unsupported format: {suffix}. Use .xlsx or .xls"}))
    except Exception as e:
        print(json.dumps({"error": str(e)}))


if __name__ == "__main__":
    main()
