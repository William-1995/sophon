#!/usr/bin/env python3
"""Read Excel/CSV files (paths default to workspace/<user_id>)."""
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def _resolve_path(params: dict, file_path: str) -> Path:
    """Resolve file path. workspace_root is already workspace/<user_id>."""
    p = Path(file_path)
    if p.is_absolute():
        return p
    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    user_id = str(params.get("user_id") or "default_user")
    if p.parts and p.parts[0] == user_id:
        p = Path(*p.parts[1:]) if len(p.parts) > 1 else Path(".")
    return workspace_root / p


def read_csv(file_path: Path, limit: int | None = None, offset: int = 0) -> dict:
    """Read CSV file."""
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        data = []
        for i, row in enumerate(reader):
            if i < offset:
                continue
            if limit is not None and len(data) >= limit:
                break
            data.append(dict(row))
    return {"headers": headers, "data": data, "total_rows": len(data)}


def read_excel(file_path: Path, sheet=0, limit: int | None = None, offset: int = 0) -> dict:
    """Read Excel file (.xlsx or .xls)."""
    suffix = file_path.suffix.lower()
    
    if suffix == '.xlsx':
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        
        headers = []
        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell is not None else f"Col_{j}" for j, cell in enumerate(row)]
                continue
            if i - 1 < offset:
                continue
            if limit is not None and len(data) >= limit:
                break
            row_dict = {headers[j]: str(cell) if cell is not None else "" for j, cell in enumerate(row) if j < len(headers)}
            data.append(row_dict)
        return {"headers": headers, "data": data, "total_rows": len(data)}
    
    elif suffix == '.xls':
        if not HAS_XLRD:
            return {"error": "xlrd not installed. Run: pip install xlrd"}
        wb = xlrd.open_workbook(str(file_path))
        if isinstance(sheet, int):
            ws = wb.sheet_by_index(sheet) if sheet < wb.nsheets else wb.sheet_by_index(0)
        else:
            try:
                ws = wb.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                ws = wb.sheet_by_index(0)
        
        headers = [str(ws.cell_value(0, j)) for j in range(ws.ncols)]
        data = []
        for i in range(1 + offset, ws.nrows):
            if limit is not None and len(data) >= limit:
                break
            row_dict = {headers[j]: str(ws.cell_value(i, j)) for j in range(ws.ncols)}
            data.append(row_dict)
        return {"headers": headers, "data": data, "total_rows": len(data)}
    
    else:
        return {"error": f"Unsupported file format: {suffix}"}


def main() -> None:
    params = json.load(sys.stdin)
    args = params.get("arguments") or params
    file_path = args.get("file", "")
    sheet = args.get("sheet", params.get("sheet", 0))
    limit = args.get("limit", params.get("limit"))
    offset = args.get("offset", params.get("offset", 0))

    if not file_path:
        print(json.dumps({"error": "file parameter is required"}))
        return

    full_path = _resolve_path(params, str(file_path))
    if not full_path.exists():
        print(json.dumps({"error": f"File not found: {full_path}"}))
        return

    suffix = full_path.suffix.lower()

    try:
        if suffix == ".csv":
            result = read_csv(full_path, limit, offset)
        elif suffix in [".xlsx", ".xls"]:
            result = read_excel(full_path, sheet, limit, offset)
        else:
            result = {"error": f"Unsupported file format: {suffix}. Use .csv, .xlsx, or .xls"}

        print(json.dumps(result, ensure_ascii=False))
    except Exception as e:  # noqa: BLE001
        print(json.dumps({"error": str(e)}))


if __name__ == "__main__":
    main()
