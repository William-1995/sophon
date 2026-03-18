#!/usr/bin/env python3
"""Return sheet structure: headers and total row count (no data)."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def _resolve_path(params: dict, file_path: str) -> Path:
    p = Path(file_path)
    if p.is_absolute():
        return p
    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    user_id = str(params.get("user_id") or "default_user")
    if p.parts and p.parts[0] == user_id:
        p = Path(*p.parts[1:]) if len(p.parts) > 1 else Path(".")
    return workspace_root / p


def structure_xlsx(full_path: Path, sheet: str | int | None) -> dict:
    """Return headers and total_rows for one sheet."""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    if not full_path.exists():
        return {"error": f"File not found: {full_path}"}
    wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
        elif sheet is not None:
            if sheet not in wb.sheetnames:
                return {"error": f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"}
            ws = wb[sheet]
        else:
            ws = wb.active
        sheet_name = ws.title
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = []
        if first_row:
            for j, cell in enumerate(first_row):
                name = str(cell).strip() if cell is not None else ""
                if not name:
                    name = f"Col_{j}"
                headers.append(name)
        total_rows = max(0, ws.max_row - 1)
        return {"sheet": sheet_name, "headers": headers, "total_rows": total_rows}
    finally:
        wb.close()


def structure_xls(full_path: Path, sheet: str | int | None) -> dict:
    """Return headers and total_rows for .xls sheet."""
    if not HAS_XLRD:
        return {"error": "xlrd not installed. Run: pip install xlrd"}
    if not full_path.exists():
        return {"error": f"File not found: {full_path}"}
    wb = xlrd.open_workbook(str(full_path))
    if isinstance(sheet, int):
        ws = wb.sheet_by_index(sheet) if sheet < wb.nsheets else wb.sheet_by_index(0)
    elif sheet is not None:
        try:
            ws = wb.sheet_by_name(str(sheet))
        except Exception:
            return {"error": f"Sheet '{sheet}' not found"}
    else:
        ws = wb.sheet_by_index(0)
    sheet_name = ws.name
    headers = [str(ws.cell_value(0, j)) or f"Col_{j}" for j in range(ws.ncols)]
    total_rows = max(0, ws.nrows - 1)
    return {"sheet": sheet_name, "headers": headers, "total_rows": total_rows}


def main() -> None:
    params = json.load(sys.stdin)
    args = params.get("arguments") or params
    file_path = args.get("file", "").strip()
    sheet = args.get("sheet", args.get("sheet_name"))

    if not file_path:
        print(json.dumps({"error": "file parameter is required"}))
        return

    full_path = _resolve_path(params, file_path)
    suffix = full_path.suffix.lower()

    try:
        if suffix == ".xlsx":
            result = structure_xlsx(full_path, sheet)
        elif suffix == ".xls":
            result = structure_xls(full_path, sheet)
        elif suffix == ".csv":
            result = {"error": "structure not supported for CSV; use read with limit=1"}
        else:
            result = {"error": f"Unsupported format: {suffix}"}
        print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"error": str(e)}))


if __name__ == "__main__":
    main()
