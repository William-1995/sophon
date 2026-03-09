"""
Excel-Ops Enrich - Excel operations.

Provides Excel file loading, header extraction, and cell operations.
"""

from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover - environment-specific
    openpyxl = None  # type: ignore


def load_sheet(full_path: Path, sheet_name: str | None) -> tuple[tuple[Any, Any] | None, dict[str, str]]:
    """Load Excel workbook and sheet.

    Args:
        full_path: Path to Excel file.
        sheet_name: Name of sheet to load, or None for active sheet.

    Returns:
        Tuple of ((workbook, worksheet), error_dict).
        If successful, returns ((wb, ws), {}).
        If failed, returns (None, {"error": "..."}).
    """
    if openpyxl is None:
        return None, {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not full_path.exists():
        return None, {"error": f"File not found: {full_path}"}

    if full_path.suffix.lower() != ".xlsx":
        return None, {"error": f"Only .xlsx is supported for enrich, got {full_path.suffix}"}

    wb = openpyxl.load_workbook(full_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return None, {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
        ws = wb[sheet_name]
    else:
        ws = wb.active

    return (wb, ws), {}


def extract_headers(ws) -> list[str]:
    """Extract header names from first row of worksheet.

    Args:
        ws: openpyxl worksheet object.

    Returns:
        List of header names. Empty cells get auto-generated names like "Col_0".
    """
    headers: list[str] = []
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if not first:
        return headers

    for idx, cell in enumerate(first):
        name = str(cell).strip() if cell is not None else ""
        if not name:
            name = f"Col_{idx}"
        headers.append(name)

    return headers


def build_column_index(headers: list[str]) -> dict[str, int]:
    """Build column name to index mapping (1-based).

    Args:
        headers: List of header names.

    Returns:
        Dict mapping header name to column index (1-based).
    """
    return {name: idx + 1 for idx, name in enumerate(headers)}


def ensure_columns_exist(
    ws,
    headers: list[str],
    col_index: dict[str, int],
    columns: list[str],
) -> tuple[list[str], dict[str, int]]:
    """Ensure columns exist in worksheet, adding them if necessary.

    Args:
        ws: openpyxl worksheet object.
        headers: Current headers list.
        col_index: Current column index mapping.
        columns: Columns to ensure exist.

    Returns:
        Updated (headers, col_index) after adding any missing columns.
    """
    headers = list(headers)  # Make a copy
    col_index = dict(col_index)  # Make a copy

    for col in columns:
        if col not in col_index:
            headers.append(col)
            col_index[col] = len(headers)
            ws.cell(row=1, column=col_index[col], value=col)

    return headers, col_index
