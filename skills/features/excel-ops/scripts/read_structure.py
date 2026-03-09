#!/usr/bin/env python3
"""
excel-ops/read_structure action.

Returns one sheet's headers and total row count (no data). LLM uses this to understand layout.
"""
import json
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(message)s", stream=sys.stderr)

_PROJECT_ROOT = Path(__file__).resolve().parents[4]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _resolve_path(params: dict, path_raw: str) -> Path:
    p = Path(path_raw.strip())
    if p.is_absolute():
        return p
    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    user_id = str(params.get("user_id") or "default_user")
    if p.parts and p.parts[0] == user_id:
        p = Path(*p.parts[1:]) if len(p.parts) > 1 else Path(".")
    return (workspace_root / p).resolve()


def _read_structure(full_path: Path, sheet_name: str | None) -> dict:
    if openpyxl is None:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    if not full_path.exists():
        return {"error": f"File not found: {full_path}"}
    if full_path.suffix.lower() != ".xlsx":
        return {"error": "Only .xlsx is supported"}

    wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = []
        if first_row:
            for j, cell in enumerate(first_row):
                name = str(cell).strip() if cell is not None else ""
                if not name:
                    name = f"Col_{j}"
                headers.append(name)
        total_rows = ws.max_row
        return {
            "sheet": sheet_name,
            "headers": headers,
            "total_rows": max(0, total_rows - 1),
        }
    finally:
        wb.close()


def main() -> None:
    params = json.loads(sys.stdin.read())
    args = params.get("arguments") or params
    path_raw = str(args.get("path", "")).strip()
    sheet_name = args.get("sheet_name") or args.get("sheet")

    if not path_raw:
        logger.warning("[excel-ops.read_structure] missing path")
        print(json.dumps({"error": "path is required"}), flush=True)
        return

    full_path = _resolve_path(params, path_raw)
    result = _read_structure(full_path, sheet_name)
    if "error" not in result:
        logger.info(
            "[excel-ops.read_structure] path=%s sheet=%s headers_count=%d total_rows=%d",
            path_raw,
            result.get("sheet", ""),
            len(result.get("headers", [])),
            result.get("total_rows", 0),
        )
    print(json.dumps(result, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
