#!/usr/bin/env python3
"""
excel-ops/enrich action.

Reads Excel sheet(s), uses id_column (and optional company_column) for search per row.
Uses search skill only (no crawl). Batches rows: N searches in parallel -> 1 LLM call -> write.
Copies copy_columns from source to target when cross-sheet. Writes to *_FILLED.xlsx by default.
"""
import asyncio
import json
import logging
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

# Ensure only JSON goes to stdout (executor parses it). Redirect all logging to stderr.
logging.basicConfig(level=logging.INFO, format="%(message)s", stream=sys.stderr)

# Ensure we can import core.* when run as a subprocess (project_root first for core)
_project_root = Path(
    os.environ.get("SOPHON_ROOT", Path(__file__).resolve().parent.parent.parent.parent.parent)
)
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))
# Load skill constants explicitly (core uses project constants)
import importlib.util
_skill_constants_path = Path(__file__).resolve().parent.parent / "constants.py"
_spec = importlib.util.spec_from_file_location("excel_ops_constants", _skill_constants_path)
_skill_constants = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_skill_constants)
DB_FILENAME = _skill_constants.DB_FILENAME  # type: ignore
from core.executor import execute_skill  # type: ignore
from providers import get_provider  # type: ignore

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover - environment-specific
    openpyxl = None  # type: ignore


@dataclass
class ResolvedContext:
    workspace_root: Path
    session_id: str
    user_id: str
    db_path: Path
    call_stack: list[str]


def _resolve_context(params: dict) -> ResolvedContext:
    workspace_root = Path(params.get("workspace_root") or ".").resolve()
    session_id = params.get("_executor_session_id") or params.get("session_id") or "default"
    session_id = str(session_id)
    user_id = params.get("user_id") or "default_user"
    user_id = str(user_id)
    db_path_raw = params.get("db_path")
    db_path = Path(db_path_raw) if db_path_raw else workspace_root / DB_FILENAME
    call_stack = list(params.get("_call_stack") or [])
    return ResolvedContext(
        workspace_root=workspace_root,
        session_id=session_id,
        user_id=user_id,
        db_path=db_path,
        call_stack=call_stack,
    )


def _load_sheet(full_path: Path, sheet_name: str | None):
    if openpyxl is None:
        return None, {"error": "openpyxl not installed. Run: pip install openpyxl"}
    if not full_path.exists():
        return None, {"error": f"File not found: {full_path}"}
    if full_path.suffix.lower() != ".xlsx":
        return None, {"error": f"Only .xlsx is supported for enrich, got {full_path.suffix}"}
    wb = openpyxl.load_workbook(full_path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return None, {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
        ws = wb[sheet_name]
    else:
        ws = wb.active
    return (wb, ws), {}


def _extract_headers(ws) -> list[str]:
    """Return header names from first row."""
    headers: list[str] = []
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first:
        return headers
    for idx, cell in enumerate(first):
        name = str(cell).strip() if cell is not None else ""
        if not name:
            name = f"Col_{idx}"
        headers.append(name)
    return headers


def _build_search_query(row_values: dict[str, Any], id_column: str, company_column: str | None) -> str:
    """Build search query: company + domain/url when company_column set, else id_column value."""
    id_val = str(row_values.get(id_column) or "").strip()
    if not id_val:
        return ""
    if company_column:
        company_val = str(row_values.get(company_column) or "").strip()
        if company_val:
            return f"{company_val} {id_val}"
    return id_val


def _extract_json(text: str) -> str:
    """Extract JSON from markdown code fences or raw text."""
    s = str(text).strip()
    # Trim markdown code fences
    if "```" in s:
        m = re.search(r"```(?:json)?\s*([\s\S]*?)```", s)
        if m:
            return m.group(1).strip()
        first = s.find("```")
        rest = s[first + 3 :]
        if "```" in rest:
            inner = rest.split("```", 1)[0]
            return inner.strip()
    return s


async def _enrich_async(params: dict) -> dict[str, Any]:
    args = params.get("arguments") or params

    path_raw = str(args.get("path", "")).strip()
    sheet_name = args.get("sheet_name")
    source_sheet = args.get("source_sheet") or sheet_name
    target_sheet = args.get("target_sheet") or source_sheet
    id_column = str(args.get("id_column", "")).strip()
    company_column = args.get("company_column")  # optional: combine with id for search
    target_columns = args.get("target_columns") or []
    copy_columns = args.get("copy_columns") or []  # copy from source to target
    start_row = int(args.get("start_row", 2))
    max_rows = args.get("max_rows")
    batch_size = int(args.get("batch_size", 5))  # rows per LLM batch (default 5)
    mode = str(args.get("mode", "copy")).strip().lower() or "copy"
    instructions = str(args.get("instructions", "")).strip()

    if not path_raw:
        return {"error": "path is required"}
    if not id_column:
        return {"error": "id_column is required"}
    if not isinstance(target_columns, list) or not target_columns:
        return {"error": "target_columns must be a non-empty array of column names"}

    ctx = _resolve_context(params)
    workspace_root = ctx.workspace_root
    raw_path = Path(path_raw)
    if raw_path.is_absolute():
        full_path = raw_path
    else:
        if raw_path.parts and raw_path.parts[0] == ctx.user_id:
            raw_path = Path(*raw_path.parts[1:]) if len(raw_path.parts) > 1 else Path(".")
        full_path = (workspace_root / raw_path).resolve()

    sheet_tuple, err = _load_sheet(full_path, source_sheet)
    if err:
        return err
    wb, ws_source = sheet_tuple  # type: ignore[misc]

    cross_sheet = source_sheet != target_sheet
    if cross_sheet:
        if target_sheet not in wb.sheetnames:
            return {"error": f"target_sheet '{target_sheet}' not found. Available: {wb.sheetnames}"}
        ws_target = wb[target_sheet]
    else:
        ws_target = ws_source

    headers_source = _extract_headers(ws_source)
    if not headers_source:
        return {"error": "Source sheet header row (row 1) is empty"}
    col_index_source: dict[str, int] = {name: idx + 1 for idx, name in enumerate(headers_source)}

    if id_column not in col_index_source:
        return {"error": f"id_column '{id_column}' not found in source headers: {headers_source}"}
    if company_column and company_column not in col_index_source:
        return {"error": f"company_column '{company_column}' not found in source headers: {headers_source}"}

    headers_target = _extract_headers(ws_target)
    if not headers_target:
        return {"error": "Target sheet header row (row 1) is empty"}
    col_index_target: dict[str, int] = {name: idx + 1 for idx, name in enumerate(headers_target)}

    # Ensure target columns exist; add copy_columns to target if missing
    for col in list(target_columns) + list(copy_columns or []):
        if col not in col_index_target:
            headers_target.append(col)
            col_index_target[col] = len(headers_target)
            ws_target.cell(row=1, column=col_index_target[col], value=col)

    max_row = ws_source.max_row
    if max_rows is not None:
        try:
            max_rows_int = int(max_rows)
        except (TypeError, ValueError):
            return {"error": f"Invalid max_rows: {max_rows!r}"}
        end_row = min(start_row + max_rows_int - 1, max_row)
    else:
        end_row = max_row
    if start_row < 2:
        start_row = 2
    if start_row > end_row:
        return {"error": f"Empty range: start_row={start_row}, end_row={end_row}"}

    provider = get_provider()

    def _read_row(row_idx: int) -> dict[str, Any]:
        row_values: dict[str, Any] = {}
        for name, idx in col_index_source.items():
            cell = ws_source.cell(row=row_idx, column=idx)
            row_values[name] = cell.value
        return row_values

    def _copy_row_to_target(row_idx: int, row_values: dict[str, Any]) -> None:
        """Copy copy_columns from source row to target row."""
        if not copy_columns:
            return
        for col in copy_columns:
            if col not in col_index_source or col not in col_index_target:
                continue
            val = row_values.get(col)
            if val is None:
                val = ""
            ws_target.cell(row=row_idx, column=col_index_target[col], value=val)

    async def _search_row(row_values: dict[str, Any]) -> str:
        query = _build_search_query(row_values, id_column, company_column)
        if not query:
            return ""
        search_result = await execute_skill(
            skill_name="search",
            action="search",
            arguments={"query": query, "num": 5},
            workspace_root=workspace_root,
            session_id=ctx.session_id,
            user_id=ctx.user_id,
            root=_project_root,
            db_path=ctx.db_path if ctx.db_path.exists() else None,
            call_stack=ctx.call_stack,
        )
        if search_result.get("error"):
            return ""
        return str(search_result.get("result") or "")

    async def _enrich_batch(row_indices: list[int]) -> tuple[int, list[dict[str, Any]]]:
        """Process a batch: read rows, search in parallel, one LLM call, return (updated_count, errors)."""
        rows_data: list[dict[str, Any]] = []
        for row_idx in row_indices:
            row_values = _read_row(row_idx)
            query = _build_search_query(row_values, id_column, company_column)
            if not query:
                continue
            rows_data.append({"row_idx": row_idx, "row_values": row_values, "query": query})

        if not rows_data:
            return 0, []

        # Parallel search for all rows in batch
        search_results = await asyncio.gather(
            *[_search_row(r["row_values"]) for r in rows_data],
            return_exceptions=True,
        )

        for r, content in zip(rows_data, search_results):
            r["search_content"] = str(content) if not isinstance(content, Exception) else ""

        # Copy copy_columns from source to target for ALL rows in batch (including empty-id rows)
        for row_idx in row_indices:
            _copy_row_to_target(row_idx, _read_row(row_idx))

        # Build batch prompt for LLM
        batch_parts = []
        for r in rows_data:
            batch_parts.append(
                f"Row {r['row_idx']} (id={r['query']}):\n"
                f"  Excel row: {json.dumps(r['row_values'], ensure_ascii=False)}\n"
                f"  Search snippets:\n{r['search_content'][:8000]}\n"
            )

        sys_prompt = (
            "You are an assistant that fills Excel columns from web search snippets. "
            "Return ONLY a JSON array of objects. Each object corresponds to one row in order. "
            "Each object must have keys for every target column. Use empty string when uncertain."
        )
        user_instr = instructions or (
            f"Infer values for target columns {target_columns} from the search snippets. "
            "If unsure, use empty string."
        )
        user_prompt = (
            f"Target columns: {target_columns}\n"
            f"Instructions: {user_instr}\n\n"
            "Batch (rows in order):\n"
            + "\n---\n".join(batch_parts)
            + "\n\n"
            "Respond ONLY with a JSON array of objects, one per row, each object mapping target column names to values. "
            "Example: [{\"col1\": \"val1\", \"col2\": \"\"}, ...]"
        )

        resp = await provider.chat(
            [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_prompt},
            ]
        )
        content = resp.get("content") or ""
        text = _extract_json(content)

        try:
            arr = json.loads(text)
        except Exception as exc:  # noqa: BLE001
            return 0, [{"row": r["row_idx"], "id_value": r["query"], "reason": f"LLM JSON parse error: {exc!s}"} for r in rows_data]

        if not isinstance(arr, list):
            return 0, [{"row": rows_data[0]["row_idx"], "reason": f"LLM output is not an array: {type(arr)!r}"}]

        rows_updated = 0
        errors: list[dict[str, Any]] = []
        for i, r in enumerate(rows_data):
            mapping = arr[i] if i < len(arr) and isinstance(arr[i], dict) else {}
            row_idx = r["row_idx"]
            if not mapping:
                errors.append({"row": row_idx, "id_value": r["query"], "reason": "No mapping in LLM response"})
                continue
            row_updated = False
            for col in target_columns:
                if col not in col_index_target:
                    continue
                value = mapping.get(col)
                if value is None:
                    value = ""
                ws_target.cell(row=row_idx, column=col_index_target[col], value=value)
                row_updated = True
            if row_updated:
                rows_updated += 1
            logger.info("[excel-ops.enrich] row=%d batch_updated", row_idx)

        return rows_updated, errors

    processed_rows = 0
    updated_rows = 0
    errors: list[dict[str, Any]] = []

    current = start_row
    while current <= end_row:
        batch_end = min(current + batch_size - 1, end_row)
        row_indices = list(range(current, batch_end + 1))
        batch_updated, batch_errors = await _enrich_batch(row_indices)
        processed_rows += len(row_indices)
        updated_rows += batch_updated
        for e in batch_errors:
            errors.append(e)
        current = batch_end + 1

    if mode == "overwrite":
        out_path = full_path
    else:
        stem = full_path.stem
        suffix = full_path.suffix
        out_name = f"{stem}_ENRICHED{suffix}" if stem.endswith("_FILLED") else f"{stem}_FILLED{suffix}"
        out_path = full_path.with_name(out_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    result: dict[str, Any] = {
        "output_path": str(out_path.relative_to(workspace_root)) if str(out_path).startswith(str(workspace_root)) else str(out_path),
        "processed_rows": processed_rows,
        "updated_rows": updated_rows,
        "target_columns": target_columns,
    }
    if copy_columns:
        result["copy_columns"] = copy_columns
    if errors:
        result["errors"] = errors
    logger.info("[excel-ops.enrich] done processed=%d errors=%d output=%s", processed_rows, len(errors), result.get("output_path", ""))
    return result


def main() -> None:
    params = json.loads(sys.stdin.read())
    result = asyncio.run(_enrich_async(params))
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
